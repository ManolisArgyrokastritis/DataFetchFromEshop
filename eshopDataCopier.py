import time
import random
import logging
import pandas as pd
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, σαν το Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, σαν το Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, σαν το Gecko) Chrome/91.0.4472.124 Safari/537.36",
]

def fetch_eshop_page(url, proxy=None):
    options = Options()
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument(f"user-agent={random.choice(user_agents)}")
    
    if proxy:
        options.add_argument(f'--proxy-server={proxy}')
    
    driver = None
    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        driver.get(url)
        driver.implicitly_wait(10)
        logging.info("Navigating to the page...")
        time.sleep(5)  # Allow page to fully load
    except WebDriverException as e:
        logging.error(f"An error occurred while setting up the driver: {e}")
        if driver:
            driver.quit()
        return None
    return driver

def click_element_by_id(driver, element_id):
    try:
        logging.info(f"Attempting to click element with id: {element_id}")
        element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, element_id))
        )
        element.click()
        logging.info(f"Clicked element with id: {element_id}")
    except TimeoutException:
        logging.error(f"Timed out waiting for the element with id: {element_id} to be clickable.")
    except NoSuchElementException:
        logging.error(f"No such element with id: {element_id}")
    except Exception as e:
        logging.error(f"An error occurred while clicking element with id: {element_id}: {e}")

def click_element_by_alt_text(driver, alt_text):
    try:
        logging.info(f"Attempting to click element with alt text: {alt_text}")
        element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, f'//img[@alt="{alt_text}"]'))
        )
        element.click()
        logging.info(f"Clicked element with alt text: {alt_text}")
    except TimeoutException:
        logging.error(f"Timed out waiting for the element with alt text: {alt_text} to be clickable.")
    except NoSuchElementException:
        logging.error(f"No such element with alt text: {alt_text}")
    except Exception as e:
        logging.error(f"An error occurred while clicking element with alt text: {alt_text}: {e}")

def click_element_by_class_and_text(driver, class_name, text):
    try:
        logging.info(f"Attempting to click element with class: {class_name} and text: {text}")
        element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, f'//h2[contains(@class, "{class_name}") and contains(text(), "{text}")]'))
        )
        element.click()
        logging.info(f"Clicked element with class: {class_name} and text: {text}")
    except TimeoutException:
        logging.error(f"Timed out waiting for the element with class: {class_name} and text: {text} to be clickable.")
    except NoSuchElementException:
        logging.error(f"No such element with class: {class_name} and text: {text}")
    except Exception as e:
        logging.error(f"An error occurred while clicking element with class: {class_name} and text: {text}: {e}")

def click_element_by_text(driver, text):
    try:
        logging.info(f"Attempting to click element with text: {text}")
        element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, f'//*[text()="{text}"]'))
        )
        element.click()
        logging.info(f"Clicked element with text: {text}")
    except TimeoutException:
        logging.error(f"Timed out waiting for the element with text: {text} to be clickable.")
    except NoSuchElementException:
        logging.error(f"No such element with text: {text}")
    except Exception as e:
        logging.error(f"An error occurred while clicking element with text: {text}: {e}")

def human_like_scroll(driver, direction="down"):
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        scroll_pause_time = random.uniform(2, 5)
        if direction == "down":
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        else:
            driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(scroll_pause_time)
        new_height = driver.execute_script("return document.body.scrollHeight")
        logging.info(f"Scrolled to {direction}, new height: {new_height}, last height: {last_height}")
        if new_height == last_height:
            break
        last_height = new_height

def extract_product_info(driver):
    product_list = []
    try:
        while True:
            logging.info("Waiting for the product container to be present...")
            WebDriverWait(driver, 60).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, '.products'))
            )
            logging.info("Product container is present.")
            
            product_container = driver.find_element(By.CSS_SELECTOR, '.products')
            products = product_container.find_elements(By.CSS_SELECTOR, '.product')
            logging.info(f"Found {len(products)} product elements")
            
            for index, product in enumerate(products):
                try:
                    logging.info(f"Extracting product {index + 1}")
                    name = product.find_element(By.CSS_SELECTOR, '.woocommerce-loop-product__title').text.strip()
                    product_list.append({'Name': name})
                except Exception as e:
                    logging.error(f"Error processing product at index {index}: {e}")
                    continue
            
            # Save products to Excel after each page
            save_to_excel(product_list, append=True)
            product_list.clear()
            
            try:
                next_page_element = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.page-numbers.next'))
                )
                next_page_url = next_page_element.get_attribute('href')
                logging.info(f"Navigating to the next page: {next_page_url}")
                driver.get(next_page_url)
                time.sleep(5)  # Allow page to fully load
            except TimeoutException:
                logging.info("No more pages found. Ending pagination.")
                break
            except NoSuchElementException:
                logging.info("No more pages found. Ending pagination.")
                break
        
    except TimeoutException:
        logging.error("Timed out waiting for the product container to load.")
    except NoSuchElementException as e:
        logging.error(f"Error extracting product information: {e}")
    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")

def save_to_excel(product_list, filename="product_names.xlsx", append=False):
    try:
        if append:
            # Load existing workbook or create a new one if it doesn't exist
            try:
                workbook = load_workbook(filename)
                sheet = workbook.active
            except FileNotFoundError:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Name"])  # Adding header if creating new file
            
            # Add data to sheet
            for product in product_list:
                sheet.append([product['Name']])
            
            workbook.save(filename)
        else:
            df = pd.DataFrame(product_list)
            df.to_excel(filename, index=False)
        
        logging.info("Product names saved to product_names.xlsx")
    except Exception as e:
        logging.error(f"Error saving to Excel: {e}")

def main(url, element_id, alt_text, class_name, text, additional_text, proxy=None):
    driver = fetch_eshop_page(url, proxy=proxy)
    if driver is None:
        logging.error("Failed to set up the webdriver. Exiting.")
        return
    
    logging.info("Page fetched successfully.")
    
    # Click the element with the specified id
    click_element_by_id(driver, element_id)
    
    # Wait for the new page to load
    time.sleep(2)  # Adjust the sleep time if necessary
    
    # Click the element with the specified alt text
    click_element_by_alt_text(driver, alt_text)
    
    # Wait for the new page to load
    time.sleep(2)  # Adjust the sleep time if necessary
    
    # Click the element with the specified class and text
    click_element_by_class_and_text(driver, class_name, text)
    
    # Wait for the new page to load
    time.sleep(2)  # Adjust the sleep time if necessary
    
    # Click the element with the specified text
    click_element_by_text(driver, additional_text)
    
    # Wait for the new page to load
    time.sleep(2)  # Adjust the sleep time if necessary
    
    # Extract product information from the new page
    extract_product_info(driver)
    
    driver.quit()

url = 'https://delovi.biz/'
element_id = 'menu-item-570'  # The id of the first element to click
alt_text = 'Elektrika'  # The alt text of the second element to click
class_name = 'woocommerce-loop-category__title'  # The class name of the third element to click
text = 'Elektrika'  # The text contained in the third element to click
additional_text = 'Alternatori - Anlaseri - Delovi'  # The text of the fourth element to click

if __name__ == "__main__":
    main(url, element_id, alt_text, class_name, text, additional_text)
